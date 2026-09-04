from __future__ import annotations

import unicodedata
from datetime import datetime
from pathlib import Path
from typing import Any

import duckdb
from openpyxl import load_workbook

from src.contracts.munic import (
    MUNIC_WORKSHEET,
    OUTPUT_STATUS_FIELDS,
    PLANNING_FIELDS,
    SOURCE_COLUMNS,
)


def _plain(value: str) -> str:
    return "".join(
        character
        for character in unicodedata.normalize("NFKD", value)
        if not unicodedata.combining(character)
    ).casefold()


def normalize_response(value: Any) -> str:
    normalized = _plain(str(value).strip()) if value is not None else ""
    mapping = {
        "sim": "declared_yes",
        "nao": "declared_no",
        "recusa": "refused",
        "nao informou": "not_reported",
        "-": "not_applicable",
    }
    if normalized not in mapping:
        raise ValueError(f"Resposta MUNIC nao reconhecida: {value!r}")
    return mapping[normalized]


def aggregate_status(statuses: list[str]) -> str:
    if "declared_yes" in statuses:
        return "declared_yes"
    if statuses and all(status == "declared_no" for status in statuses):
        return "declared_no"
    if "refused" in statuses:
        return "refused"
    if "not_reported" in statuses:
        return "not_reported"
    if "unknown" in statuses:
        return "unknown"
    return "not_applicable"


def load_munic_records(
    workbook_path: Path,
    *,
    source: str,
    source_url: str,
    reference_year: int,
    ingested_at: datetime,
) -> list[dict[str, Any]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        worksheet = next(
            (
                sheet
                for sheet in workbook.worksheets
                if _plain(sheet.title) == _plain(MUNIC_WORKSHEET)
            ),
            None,
        )
        if worksheet is None:
            raise ValueError(f"Aba {MUNIC_WORKSHEET!r} nao encontrada na MUNIC")

        rows = worksheet.iter_rows(values_only=True)
        headers = [str(value).strip() if value is not None else "" for value in next(rows)]
        header_index = {header.casefold(): index for index, header in enumerate(headers)}
        missing = [
            source_column
            for source_column in SOURCE_COLUMNS.values()
            if source_column.casefold() not in header_index
        ]
        if missing:
            raise ValueError(f"Colunas esperadas ausentes na MUNIC: {missing}")

        def value(row: tuple[Any, ...], source_column: str) -> Any:
            return row[header_index[source_column.casefold()]]

        records: list[dict[str, Any]] = []
        for row in rows:
            codigo_ibge = str(value(row, SOURCE_COLUMNS["codigo_ibge"])).strip()
            planning_statuses = {
                f"{field}_status": normalize_response(
                    value(row, SOURCE_COLUMNS[field])
                )
                for field in (
                    "flood_planning_master_plan",
                    "flood_planning_land_use_law",
                    "flood_planning_specific_law",
                    "landslide_planning_master_plan",
                    "landslide_planning_land_use_law",
                    "landslide_planning_specific_law",
                )
            }
            civil_defense_status = normalize_response(
                value(row, SOURCE_COLUMNS["municipal_civil_defense_body"])
            )
            if normalize_response(
                value(row, SOURCE_COLUMNS["municipal_civil_defense_unknown"])
            ) == "declared_yes":
                civil_defense_status = "unknown"

            record = {
                "codigo_ibge": codigo_ibge,
                "sigla_uf_fonte": str(value(row, SOURCE_COLUMNS["sigla_uf"])).strip(),
                "municipio_fonte": str(
                    value(row, SOURCE_COLUMNS["municipio_fonte"])
                ).strip(),
                **planning_statuses,
                "any_risk_prevention_planning_instrument_status": aggregate_status(
                    [planning_statuses[field] for field in PLANNING_FIELDS]
                ),
                "flood_risk_mapping_status": normalize_response(
                    value(row, SOURCE_COLUMNS["flood_risk_mapping"])
                ),
                "flood_contingency_plan_status": normalize_response(
                    value(row, SOURCE_COLUMNS["flood_contingency_plan"])
                ),
                "flood_early_warning_status": normalize_response(
                    value(row, SOURCE_COLUMNS["flood_early_warning"])
                ),
                "landslide_risk_mapping_status": normalize_response(
                    value(row, SOURCE_COLUMNS["landslide_risk_mapping"])
                ),
                "landslide_contingency_plan_status": normalize_response(
                    value(row, SOURCE_COLUMNS["landslide_contingency_plan"])
                ),
                "landslide_early_warning_status": normalize_response(
                    value(row, SOURCE_COLUMNS["landslide_early_warning"])
                ),
                "municipal_civil_defense_body_status": civil_defense_status,
                "civil_defense_budget_provision_status": normalize_response(
                    value(row, SOURCE_COLUMNS["civil_defense_budget_provision"])
                ),
                "civil_defense_early_warning_status": normalize_response(
                    value(row, SOURCE_COLUMNS["civil_defense_early_warning"])
                ),
                "source_year": reference_year,
                "source": source,
                "source_url": source_url,
                "ingested_at": ingested_at.isoformat(),
            }
            records.append(record)
    finally:
        workbook.close()
    return records


def create_munic_tables(
    records: list[dict[str, Any]], *, dim_municipality_path: Path
) -> duckdb.DuckDBPyConnection:
    connection = duckdb.connect(":memory:")
    status_columns = ",\n".join(f"{field} VARCHAR" for field in OUTPUT_STATUS_FIELDS)
    connection.execute(
        f"""
        CREATE TABLE silver_munic (
            codigo_ibge VARCHAR,
            sigla_uf_fonte VARCHAR,
            municipio_fonte VARCHAR,
            {status_columns},
            source_year INTEGER,
            source VARCHAR,
            source_url VARCHAR,
            ingested_at TIMESTAMP WITH TIME ZONE
        )
        """
    )
    columns = [
        "codigo_ibge",
        "sigla_uf_fonte",
        "municipio_fonte",
        *OUTPUT_STATUS_FIELDS,
        "source_year",
        "source",
        "source_url",
        "ingested_at",
    ]
    placeholders = ", ".join("?" for _ in columns)
    connection.executemany(
        f"INSERT INTO silver_munic VALUES ({placeholders})",
        [[record[column] for column in columns] for record in records],
    )
    dimension_path = str(dim_municipality_path).replace("'", "''")
    status_projection = ",\n".join(
        f"COALESCE(s.{field}, 'not_in_source') AS {field}"
        for field in OUTPUT_STATUS_FIELDS
    )
    connection.execute(
        f"""
        CREATE TABLE gold_munic AS
        SELECT
            d.codigo_ibge,
            d.municipio,
            d.sigla_uf,
            s.municipio_fonte,
            s.codigo_ibge IS NOT NULL AS in_source,
            {status_projection},
            COALESCE(s.source_year, 2020) AS source_year,
            s.source,
            s.source_url,
            s.ingested_at
        FROM read_parquet('{dimension_path}') d
        LEFT JOIN silver_munic s USING (codigo_ibge)
        ORDER BY d.codigo_ibge
        """
    )
    return connection


def write_munic_artifacts(
    connection: duckdb.DuckDBPyConnection,
    *,
    silver_path: Path,
    gold_path: Path,
) -> None:
    silver_path.parent.mkdir(parents=True, exist_ok=True)
    gold_path.parent.mkdir(parents=True, exist_ok=True)
    silver_sql_path = str(silver_path).replace("'", "''")
    gold_sql_path = str(gold_path).replace("'", "''")
    connection.execute(
        f"COPY silver_munic TO '{silver_sql_path}.tmp' (FORMAT PARQUET, COMPRESSION ZSTD)"
    )
    connection.execute(
        f"COPY gold_munic TO '{gold_sql_path}.tmp' (FORMAT PARQUET, COMPRESSION ZSTD)"
    )
    Path(f"{silver_path}.tmp").replace(silver_path)
    Path(f"{gold_path}.tmp").replace(gold_path)
